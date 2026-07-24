#!/usr/bin/env python3
"""Validate a browser-generated xlsx workbook against a committed expectation set.

Usage (prefer the Makefile):
  make test-browser-xlsx local_libs=1

The e2e spec downloads the real workbook the web app writes with
write-excel-file; this script loads it with openpyxl and checks its structure:
a single worksheet with the expected name, the exact header row, the data row
count, representative synthetic cell values, OWASP formula-injection escaping,
and XML-safe output. Diagnostics are content-safe: only fixed check names,
row/column ordinals, and control-character codepoints are printed. Cell values,
which echo user content in the real flow, are never printed.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# OWASP formula-injection prefixes (= + - @ TAB CR LF). A cleaned cell that
# started with one of these is quote-prefixed, so no exported cell may begin
# with them. Mirrors _FORMULA_PREFIXES in core/text.py and field-cleaners.js.
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")

# Control characters XML 1.0 forbids in a worksheet cell. They are stripped
# before export, so no exported cell may contain them. Mirrors
# _ILLEGAL_XML_CHARS_RE in core/text.py and ILLEGAL_XML_CHARS in field-cleaners.js.
ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def load_expectations(path: Path) -> dict[str, Any]:
    """Load and shape-check the committed expectation set."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("expectations must be a JSON object")
    for key in ("sheetName", "headers", "dataRowCount", "cells"):
        if key not in payload:
            raise ValueError(f"expectations missing key: {key}")
    return payload


def read_workbook(path: Path) -> tuple[list[str], list[str], list[list[str]]]:
    """Read worksheet names, the header row, and data rows as string cells."""
    with warnings.catch_warnings():
        # write-excel-file omits the default cell style openpyxl expects; the
        # substituted default is irrelevant to the value-level checks here.
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_names = list(workbook.sheetnames)
        worksheet = workbook.worksheets[0]
        rows: list[list[str]] = [
            ["" if value is None else str(value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()

    header = rows[0] if rows else []
    data_rows = rows[1:]
    return sheet_names, header, data_rows


def check_structure(
    sheet_names: list[str],
    header: list[str],
    data_rows: list[list[str]],
    expected: dict[str, Any],
) -> list[str]:
    """Check worksheet count, sheet name, header row, and data row count."""
    failures: list[str] = []

    if len(sheet_names) != 1:
        failures.append(f"SHEETS   FAIL expected=1 actual={len(sheet_names)}")
    elif sheet_names[0] != expected["sheetName"]:
        failures.append("SHEET    FAIL name-mismatch")

    expected_headers = list(expected["headers"])
    if len(header) != len(expected_headers):
        failures.append(
            f"HEADER   FAIL check=length expected-cols={len(expected_headers)} actual-cols={len(header)}"
        )
    elif header != expected_headers:
        first_mismatch = next(
            index
            for index, (actual, want) in enumerate(zip(header, expected_headers), start=1)
            if actual != want
        )
        failures.append(f"HEADER   FAIL check=values first-mismatch-column={first_mismatch}")

    if len(data_rows) != expected["dataRowCount"]:
        failures.append(
            f"ROWS     FAIL expected={expected['dataRowCount']} actual={len(data_rows)}"
        )

    return failures


def check_cells(data_rows: list[list[str]], expected: dict[str, Any]) -> list[str]:
    """Check representative synthetic cell values without printing them."""
    failures: list[str] = []
    for cell in expected["cells"]:
        row = cell["row"]
        column = cell["column"]
        if row < 1 or row > len(data_rows) or column < 1 or column > len(data_rows[row - 1]):
            failures.append(f"CELL     FAIL row={row} column={column} check=out-of-range")
            continue
        if data_rows[row - 1][column - 1] != cell["equals"]:
            failures.append(f"CELL     FAIL row={row} column={column} check=equals")
    return failures


def check_safety(data_rows: list[list[str]]) -> list[str]:
    """Check every data cell is XML-safe and free of formula-injection prefixes."""
    failures: list[str] = []
    for row_index, row in enumerate(data_rows, start=1):
        for column_index, value in enumerate(row, start=1):
            illegal = ILLEGAL_XML_CHARS_RE.search(value)
            if illegal:
                codepoint = ord(illegal.group())
                failures.append(
                    f"XML      FAIL row={row_index} column={column_index} "
                    f"codepoint=0x{codepoint:02x}"
                )
            if value and value[0] in FORMULA_PREFIXES:
                failures.append(
                    f"FORMULA  FAIL row={row_index} column={column_index} "
                    f"prefix=0x{ord(value[0]):02x}"
                )
    return failures


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse the workbook and expectation paths, with environment fallbacks."""
    parser = argparse.ArgumentParser(description=__doc__)
    # argparse applies type= only to values parsed off the command line, not to
    # environment-sourced defaults, so accept raw strings and coerce to Path
    # once below for both the CLI and env-var paths.
    parser.add_argument("--workbook", default=os.environ.get("BROWSER_XLSX_OUT"))
    parser.add_argument("--expected", default=os.environ.get("BROWSER_XLSX_EXPECTED"))
    args = parser.parse_args(argv)
    if args.workbook is None or args.expected is None:
        parser.error("both --workbook and --expected are required")
    args.workbook = Path(args.workbook)
    args.expected = Path(args.expected)
    return args


def main(argv: list[str] | None = None) -> int:
    """Validate the workbook and return nonzero on any validation failure."""
    args = parse_args(argv)

    if not args.workbook.is_file():
        print("RESULT   FAILED reason=missing-workbook")
        return 1

    try:
        expected = load_expectations(args.expected)
        sheet_names, header, data_rows = read_workbook(args.workbook)
    except Exception as error:  # noqa: BLE001 - report the failure class, not content
        print(f"RESULT   FAILED reason=load-error error={type(error).__name__}")
        return 1

    failures = check_structure(sheet_names, header, data_rows, expected)
    # check_cells guards out-of-range positions and check_safety is row-count
    # agnostic, so run both unconditionally for complete diagnostics.
    failures.extend(check_cells(data_rows, expected))
    failures.extend(check_safety(data_rows))

    for failure in failures:
        print(failure)

    if failures:
        print(f"RESULT   FAILED failures={len(failures)}")
        return 1

    print(
        f"RESULT   PASSED sheet-name-ok=1 header-ok=1 data-rows={len(data_rows)} "
        f"cell-checks={len(expected['cells'])}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
