#!/usr/bin/env python3
"""Compare Python CLI xlsx output with web cleaner JSON rows.

Usage (prefer the Makefile):
  make xrt-diff
  make xrt-diff strict=1
  make xrt-diff input_dir=/private/export xlsx_dir=/private/output

The comparison is content-safe. It reports only fixed type names, statuses,
aggregate counts, and a bounded set of row and column ordinals. Missing inputs
skip outside strict mode and fail in strict audit mode.
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
DEFAULT_XLSX_DIR = REPO / "data/output"
MAX_MISMATCH_DETAILS = 3

TYPES = {
    "shares": "Shares.xlsx",
    "comments": "Comments.xlsx",
    "messages": "Messages.xlsx",
    "connections": "Connections.xlsx",
}


@dataclass(frozen=True)
class ComparisonResult:
    """Aggregate comparison counts for one fixed data type."""

    python_rows: int
    web_rows: int
    mismatched_rows: int
    mismatched_cells: int
    extra_python_rows: int
    extra_web_rows: int
    details: tuple[tuple[int, int | None, str], ...]

    @property
    def identical(self) -> bool:
        """Return whether the two outputs contain identical rows."""
        return (
            self.mismatched_rows == 0 and self.extra_python_rows == 0 and self.extra_web_rows == 0
        )


def read_xlsx_rows(path: Path) -> list[list[Any]]:
    """Read the first xlsx worksheet into padded row value lists."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if not workbook.worksheets:
            raise ValueError("missing worksheet")
        worksheet = workbook.worksheets[0]

        rows: list[list[Any]] = []
        header_width = 0
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
            values = ["" if value is None else value for value in row]
            if row_index == 0:
                header_width = len(values)
            elif len(values) < header_width:
                values.extend([""] * (header_width - len(values)))
            rows.append(values)
        return rows
    finally:
        workbook.close()


def read_web_rows(path: Path, header: list[Any]) -> list[list[Any]]:
    """Validate and project web cleaner rows to xlsx columns in one pass."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError("invalid row payload")

    rows: list[list[Any]] = []
    for row in payload:
        if not isinstance(row, dict):
            raise ValueError("invalid row payload")
        rows.append([row.get(column, "") for column in header])
    return rows


def compare_rows(python_rows: list[list[Any]], web_rows: list[list[Any]]) -> ComparisonResult:
    """Compare all rows and retain only bounded mismatch coordinates."""
    mismatched_rows = 0
    mismatched_cells = 0
    extra_python_rows = max(len(python_rows) - len(web_rows), 0)
    extra_web_rows = max(len(web_rows) - len(python_rows), 0)
    details: list[tuple[int, int | None, str]] = []

    common_rows = min(len(python_rows), len(web_rows))
    for row_index in range(common_rows):
        python_row = python_rows[row_index]
        web_row = web_rows[row_index]
        width = max(len(python_row), len(web_row))
        row_differs = False
        for column_index in range(width):
            python_value = python_row[column_index] if column_index < len(python_row) else ""
            web_value = web_row[column_index] if column_index < len(web_row) else ""
            if python_value == web_value:
                continue
            row_differs = True
            mismatched_cells += 1
            if len(details) < MAX_MISMATCH_DETAILS:
                details.append((row_index + 1, column_index + 1, "cell"))
        if row_differs:
            mismatched_rows += 1

    remaining_details = MAX_MISMATCH_DETAILS - len(details)
    for row_index in range(common_rows, common_rows + min(extra_python_rows, remaining_details)):
        details.append((row_index + 1, None, "python-row"))

    remaining_details = MAX_MISMATCH_DETAILS - len(details)
    for row_index in range(common_rows, common_rows + min(extra_web_rows, remaining_details)):
        details.append((row_index + 1, None, "web-row"))

    return ComparisonResult(
        python_rows=len(python_rows),
        web_rows=len(web_rows),
        mismatched_rows=mismatched_rows,
        mismatched_cells=mismatched_cells,
        extra_python_rows=extra_python_rows,
        extra_web_rows=extra_web_rows,
        details=tuple(details),
    )


def compare_type(xlsx_path: Path, json_path: Path) -> ComparisonResult:
    """Load and compare one xlsx and JSON pair."""
    xlsx_rows = read_xlsx_rows(xlsx_path)
    if not xlsx_rows:
        raise ValueError("missing header row")

    header = xlsx_rows[0]
    if not header:
        raise ValueError("empty header row")

    python_rows = xlsx_rows[1:]
    web_rows = read_web_rows(json_path, header)
    return compare_rows(python_rows, web_rows)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse configurable comparison directories and strict audit mode."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx-dir", type=Path, default=DEFAULT_XLSX_DIR)
    parser.add_argument("--json-dir", type=Path, required=True)
    parser.add_argument("--strict", action="store_true")
    return parser.parse_args(argv)


def _print_result(type_name: str, result: ComparisonResult) -> None:
    status = "IDENTICAL" if result.identical else "DIFFERS"
    print(
        f"{type_name:<12} {status:<10} python-rows={result.python_rows} "
        f"web-rows={result.web_rows} mismatched-rows={result.mismatched_rows} "
        f"mismatched-cells={result.mismatched_cells} "
        f"extra-python-rows={result.extra_python_rows} extra-web-rows={result.extra_web_rows}"
    )
    for row, column, kind in result.details:
        if column is None:
            print(f"{type_name:<12} DETAIL     kind={kind} row={row}")
        else:
            print(f"{type_name:<12} DETAIL     kind={kind} row={row} column={column}")


def main(argv: list[str] | None = None) -> int:
    """Compare every configured type and return nonzero for differences or errors."""
    args = parse_args(argv)
    missing_by_type = {
        type_name: (
            not (args.xlsx_dir / xlsx_name).is_file(),
            not (args.json_dir / f"{type_name}.json").is_file(),
        )
        for type_name, xlsx_name in TYPES.items()
    }
    missing_count = sum(sum(missing) for missing in missing_by_type.values())
    if missing_count:
        for type_name, (missing_xlsx, missing_json) in missing_by_type.items():
            if missing_xlsx or missing_json:
                print(
                    f"{type_name:<12} MISSING    python-inputs={int(missing_xlsx)} "
                    f"web-inputs={int(missing_json)}"
                )
        status = "FAILED" if args.strict else "SKIPPED"
        print(f"RESULT       {status:<10} missing-inputs={missing_count}")
        return 1 if args.strict else 0

    totals = {
        "differing": 0,
        "errors": 0,
        "mismatched_rows": 0,
        "mismatched_cells": 0,
        "extra_python_rows": 0,
        "extra_web_rows": 0,
    }
    for type_name, xlsx_name in TYPES.items():
        try:
            result = compare_type(
                args.xlsx_dir / xlsx_name,
                args.json_dir / f"{type_name}.json",
            )
        except Exception:
            totals["errors"] += 1
            print(f"{type_name:<12} ERROR      comparison-failures=1")
            continue

        _print_result(type_name, result)
        if not result.identical:
            totals["differing"] += 1
        totals["mismatched_rows"] += result.mismatched_rows
        totals["mismatched_cells"] += result.mismatched_cells
        totals["extra_python_rows"] += result.extra_python_rows
        totals["extra_web_rows"] += result.extra_web_rows

    failed = totals["differing"] > 0 or totals["errors"] > 0
    status = "FAILED" if failed else "IDENTICAL"
    print(
        f"RESULT       {status:<10} types={len(TYPES)} differing={totals['differing']} "
        f"errors={totals['errors']} mismatched-rows={totals['mismatched_rows']} "
        f"mismatched-cells={totals['mismatched_cells']} "
        f"extra-python-rows={totals['extra_python_rows']} "
        f"extra-web-rows={totals['extra_web_rows']}"
    )
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
