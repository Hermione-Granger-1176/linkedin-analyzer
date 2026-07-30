"""Excel formatting utilities."""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.styles import Alignment

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

    from linkedin_analyzer.core.types import CleanerConfig


def format_excel_worksheet(ws: Worksheet | None, config: CleanerConfig) -> None:
    """Apply configured formatting to an open worksheet.

    Args:
        ws: Worksheet to format
        config: Cleaner configuration with column settings

    Raises:
        RuntimeError: If no worksheet is available
    """
    if ws is None:
        raise RuntimeError("Failed to load active worksheet")

    for col_letter, width in config.column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wrap_columns = set(config.wrap_text_columns)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            # A save and reload used to normalize empty strings to blank cells.
            # Preserve that workbook representation without a second file pass.
            if cell.value == "":
                cell.value = None
            if cell.column in wrap_columns:
                cell.alignment = wrap_alignment

    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.alignment = header_alignment


def format_excel_output(output_path: Path, config: CleanerConfig) -> None:
    """Apply formatting to an Excel file.

    Args:
        output_path: Path to the Excel file
        config: Cleaner configuration with column settings

    Raises:
        RuntimeError: If the worksheet cannot be loaded
    """
    wb = load_workbook(output_path)
    try:
        format_excel_worksheet(wb.active, config)
        wb.save(output_path)
    finally:
        wb.close()
