"""Tests for Excel formatting utilities."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import linkedin_analyzer.core.excel as excel
from linkedin_analyzer.core.types import CleanerConfig, ColumnConfig


def test_format_excel_worksheet_applies_configured_styles(tmp_path: Path) -> None:
    """Apply widths, wrapping, and header alignment before a workbook is saved."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Name", "Message"])
    worksheet.append(["Ada", "Hello"])
    worksheet.append(["Grace", ""])
    config = CleanerConfig(
        input_path=tmp_path / "input.csv",
        output_path=tmp_path / "output.xlsx",
        columns=(
            ColumnConfig(name="Name", width=24),
            ColumnConfig(name="Message", width=50, wrap_text=True),
        ),
    )

    excel.format_excel_worksheet(worksheet, config)

    assert worksheet.column_dimensions["A"].width == 24
    assert worksheet.column_dimensions["B"].width == 50
    assert worksheet["A1"].alignment.horizontal == "center"
    assert worksheet["A1"].alignment.vertical == "center"
    assert worksheet["B2"].alignment.wrap_text is True
    assert worksheet["B2"].alignment.vertical == "top"
    assert worksheet["A2"].alignment.wrap_text is None
    assert worksheet["B3"].value is None
    assert worksheet["B3"].data_type == "n"
    assert worksheet["B3"].alignment.wrap_text is True


def test_format_excel_worksheet_raises_without_active_sheet(tmp_path: Path) -> None:
    """Reject a missing worksheet before attempting to save a workbook."""
    config = CleanerConfig(
        input_path=tmp_path / "input.csv",
        output_path=tmp_path / "output.xlsx",
        columns=(ColumnConfig(name="A"),),
    )

    with pytest.raises(RuntimeError, match="Failed to load active worksheet"):
        excel.format_excel_worksheet(None, config)


def test_format_excel_output_saves_formatted_workbook(tmp_path: Path) -> None:
    """Keep the path-based formatting API working for existing callers."""
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.active.append(["Name"])
    workbook.active.append(["Ada"])
    workbook.save(output_path)
    workbook.close()
    config = CleanerConfig(
        input_path=tmp_path / "input.csv",
        output_path=output_path,
        columns=(ColumnConfig(name="Name", width=24),),
    )

    excel.format_excel_output(output_path, config)

    formatted = load_workbook(output_path)
    try:
        assert formatted.active.column_dimensions["A"].width == 24
        assert formatted.active["A1"].alignment.horizontal == "center"
    finally:
        formatted.close()


def test_format_excel_output_raises_without_active_sheet(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Raise an error when no worksheet is available."""

    class DummyWorkbook:
        active = None
        closed = False

        def close(self) -> None:
            self.closed = True

    workbook = DummyWorkbook()

    def fake_load_workbook(_: Path) -> DummyWorkbook:
        return workbook

    monkeypatch.setattr(excel, "load_workbook", fake_load_workbook)

    config = CleanerConfig(
        input_path=tmp_path / "input.csv",
        output_path=tmp_path / "output.xlsx",
        columns=(ColumnConfig(name="A"),),
    )

    with pytest.raises(RuntimeError, match="Failed to load active worksheet"):
        excel.format_excel_output(tmp_path / "output.xlsx", config)
    assert workbook.closed
