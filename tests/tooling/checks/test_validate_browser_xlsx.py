"""Tests for the browser-xlsx workbook validator behind `make test-browser-xlsx`."""

from __future__ import annotations

import json
from typing import TYPE_CHECKING, Any

import pytest
from openpyxl import Workbook
from scripts.checks import validate_browser_xlsx as validator

if TYPE_CHECKING:
    from pathlib import Path

HEADERS = ["Date", "Link", "Message"]
DATA_ROWS = [
    ["2024-01-01", "https://example.invalid/1", "Plain comment."],
    ["2024-01-02", "https://example.invalid/2", "'=SUM(1+1)"],
]


def _expectations(**overrides: Any) -> dict[str, Any]:
    """Build a committed-style expectation set with optional overrides."""
    payload: dict[str, Any] = {
        "sheetName": "Comments",
        "headers": list(HEADERS),
        "dataRowCount": len(DATA_ROWS),
        "cells": [{"row": 1, "column": 3, "equals": "Plain comment."}],
    }
    payload.update(overrides)
    return payload


def _write_workbook(
    path: Path,
    rows: list[list[str]],
    *,
    sheet_name: str = "Comments",
    extra_sheet: bool = False,
) -> None:
    """Write a synthetic workbook that stands in for the browser download."""
    workbook = Workbook()
    default_worksheet = workbook.active
    assert default_worksheet is not None
    workbook.remove(default_worksheet)
    worksheet = workbook.create_sheet(sheet_name)
    for row in rows:
        worksheet.append(row)
    if extra_sheet:
        workbook.create_sheet("Extra")
    workbook.save(path)


def _write_expectations(path: Path, payload: dict[str, Any]) -> None:
    """Write an expectation document beside the workbook."""
    path.write_text(json.dumps(payload), encoding="utf-8")


def test_load_expectations_accepts_a_complete_document(tmp_path: Path) -> None:
    """A document with every required key loads unchanged."""
    path = tmp_path / "expected.json"
    _write_expectations(path, _expectations())

    assert validator.load_expectations(path)["sheetName"] == "Comments"


def test_load_expectations_rejects_a_non_object_document(tmp_path: Path) -> None:
    """Only a JSON object can carry the expectation keys."""
    path = tmp_path / "expected.json"
    path.write_text("[]", encoding="utf-8")

    with pytest.raises(ValueError, match="must be a JSON object"):
        validator.load_expectations(path)


def test_load_expectations_requires_every_key(tmp_path: Path) -> None:
    """A missing key names itself instead of failing later during comparison."""
    path = tmp_path / "expected.json"
    payload = _expectations()
    del payload["cells"]
    _write_expectations(path, payload)

    with pytest.raises(ValueError, match="expectations missing key: cells"):
        validator.load_expectations(path)


def test_read_workbook_splits_the_header_from_the_data_rows(tmp_path: Path) -> None:
    """String-coerce every cell and separate the header row from the data."""
    workbook_path = tmp_path / "workbook.xlsx"
    _write_workbook(workbook_path, [HEADERS, *DATA_ROWS])

    sheet_names, header, data_rows = validator.read_workbook(workbook_path)

    assert sheet_names == ["Comments"]
    assert header == HEADERS
    assert data_rows == DATA_ROWS


def test_read_workbook_handles_an_empty_worksheet(tmp_path: Path) -> None:
    """An empty download yields no header and no data rows rather than an index error."""
    workbook_path = tmp_path / "empty.xlsx"
    _write_workbook(workbook_path, [])

    sheet_names, header, data_rows = validator.read_workbook(workbook_path)

    assert sheet_names == ["Comments"]
    assert header == []
    assert data_rows == []


def test_read_workbook_coerces_blank_cells_to_empty_strings(tmp_path: Path) -> None:
    """A blank trailing cell reads as an empty string, never as None."""
    workbook_path = tmp_path / "blank.xlsx"
    _write_workbook(workbook_path, [HEADERS, ["2024-01-01", None, "text"]])

    _, _, data_rows = validator.read_workbook(workbook_path)

    assert data_rows == [["2024-01-01", "", "text"]]


def test_check_structure_accepts_a_matching_workbook() -> None:
    """A workbook that matches the expectation set reports no failures."""
    assert validator.check_structure(["Comments"], HEADERS, DATA_ROWS, _expectations()) == []


def test_check_structure_reports_extra_worksheets() -> None:
    """The export must contain exactly one worksheet."""
    failures = validator.check_structure(
        ["Comments", "Extra"],
        HEADERS,
        DATA_ROWS,
        _expectations(),
    )

    assert failures == ["SHEETS   FAIL expected=1 actual=2"]


def test_check_structure_reports_a_renamed_worksheet() -> None:
    """A single worksheet with the wrong name is a structural failure."""
    failures = validator.check_structure(["Other"], HEADERS, DATA_ROWS, _expectations())

    assert failures == ["SHEET    FAIL name-mismatch"]


def test_check_structure_reports_header_length_and_value_mismatches() -> None:
    """Report the column count first, then the first differing column."""
    short = validator.check_structure(["Comments"], ["Date"], DATA_ROWS, _expectations())
    assert short == ["HEADER   FAIL check=length expected-cols=3 actual-cols=1"]

    renamed = validator.check_structure(
        ["Comments"],
        ["Date", "URL", "Message"],
        DATA_ROWS,
        _expectations(),
    )
    assert renamed == ["HEADER   FAIL check=values first-mismatch-column=2"]


def test_check_structure_reports_a_row_count_mismatch() -> None:
    """A different data row count is reported without echoing any cell."""
    failures = validator.check_structure(
        ["Comments"],
        HEADERS,
        DATA_ROWS,
        _expectations(dataRowCount=99),
    )

    assert failures == ["ROWS     FAIL expected=99 actual=2"]


def test_check_cells_accepts_matching_positions() -> None:
    """A cell expectation that holds contributes no failure."""
    assert validator.check_cells(DATA_ROWS, _expectations()) == []


def test_check_cells_reports_out_of_range_positions() -> None:
    """Positions outside the workbook are reported instead of raising IndexError."""
    positions = [
        {"row": 0, "column": 1, "equals": "x"},
        {"row": 9, "column": 1, "equals": "x"},
        {"row": 1, "column": 0, "equals": "x"},
        {"row": 1, "column": 9, "equals": "x"},
    ]

    failures = validator.check_cells(DATA_ROWS, _expectations(cells=positions))

    assert failures == [
        "CELL     FAIL row=0 column=1 check=out-of-range",
        "CELL     FAIL row=9 column=1 check=out-of-range",
        "CELL     FAIL row=1 column=0 check=out-of-range",
        "CELL     FAIL row=1 column=9 check=out-of-range",
    ]


def test_check_cells_reports_a_value_mismatch_without_printing_it() -> None:
    """A wrong cell names its position only, never its content."""
    failures = validator.check_cells(
        DATA_ROWS,
        _expectations(cells=[{"row": 1, "column": 3, "equals": "something else"}]),
    )

    assert failures == ["CELL     FAIL row=1 column=3 check=equals"]


def test_check_safety_accepts_escaped_and_empty_cells() -> None:
    """Quote-prefixed formulas and empty cells are safe export content."""
    assert validator.check_safety([["", "'=SUM(1+1)", "plain"]]) == []


def test_check_safety_reports_illegal_xml_and_formula_prefixes() -> None:
    """Report the codepoint class of unsafe cells without echoing the value."""
    failures = validator.check_safety([["a\x01b", "=SUM(1+1)"]])

    assert failures == [
        "XML      FAIL row=1 column=1 codepoint=0x01",
        "FORMULA  FAIL row=1 column=2 prefix=0x3d",
    ]


def test_parse_args_reads_the_environment_when_flags_are_absent(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The Make target may hand both paths through the environment instead."""
    monkeypatch.setenv("BROWSER_XLSX_OUT", str(tmp_path / "workbook.xlsx"))
    monkeypatch.setenv("BROWSER_XLSX_EXPECTED", str(tmp_path / "expected.json"))

    args = validator.parse_args([])

    assert args.workbook == tmp_path / "workbook.xlsx"
    assert args.expected == tmp_path / "expected.json"


def test_parse_args_requires_both_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    """Neither path may be left unset once the environment is empty."""
    monkeypatch.delenv("BROWSER_XLSX_OUT", raising=False)
    monkeypatch.delenv("BROWSER_XLSX_EXPECTED", raising=False)

    with pytest.raises(SystemExit):
        validator.parse_args(["--workbook", "book.xlsx"])


def test_main_passes_a_matching_workbook(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    """A workbook that satisfies every check reports a content-safe pass line."""
    workbook_path = tmp_path / "workbook.xlsx"
    expected_path = tmp_path / "expected.json"
    _write_workbook(workbook_path, [HEADERS, *DATA_ROWS])
    _write_expectations(expected_path, _expectations())

    exit_code = validator.main(["--workbook", str(workbook_path), "--expected", str(expected_path)])

    assert exit_code == 0
    assert "RESULT   PASSED sheet-name-ok=1 header-ok=1 data-rows=2 cell-checks=1" in (
        capsys.readouterr().out
    )


def test_main_reports_a_missing_workbook(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    """A download that never landed fails before openpyxl is asked to read it."""
    exit_code = validator.main(
        ["--workbook", str(tmp_path / "missing.xlsx"), "--expected", str(tmp_path / "e.json")]
    )

    assert exit_code == 1
    assert "RESULT   FAILED reason=missing-workbook" in capsys.readouterr().out


def test_main_reports_a_load_error_by_class(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    """An unreadable expectation set reports its error class, not its content."""
    workbook_path = tmp_path / "workbook.xlsx"
    _write_workbook(workbook_path, [HEADERS, *DATA_ROWS])

    exit_code = validator.main(
        ["--workbook", str(workbook_path), "--expected", str(tmp_path / "missing.json")]
    )

    assert exit_code == 1
    assert "RESULT   FAILED reason=load-error error=FileNotFoundError" in capsys.readouterr().out


def test_main_reports_every_failure_and_their_count(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    """Structure, cell, and safety failures are all reported in one run."""
    workbook_path = tmp_path / "workbook.xlsx"
    expected_path = tmp_path / "expected.json"
    _write_workbook(
        workbook_path,
        [HEADERS, ["2024-01-01", "https://example.invalid/1", "=SUM(1+1)"]],
        sheet_name="Renamed",
    )
    _write_expectations(expected_path, _expectations())

    exit_code = validator.main(["--workbook", str(workbook_path), "--expected", str(expected_path)])

    output = capsys.readouterr().out
    assert exit_code == 1
    assert "SHEET    FAIL name-mismatch" in output
    assert "ROWS     FAIL expected=2 actual=1" in output
    assert "CELL     FAIL row=1 column=3 check=equals" in output
    assert "FORMULA  FAIL row=1 column=3 prefix=0x3d" in output
    assert "RESULT   FAILED failures=4" in output
